from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "mtn_h3_source_map.json"
EXPORT_DIR = ROOT / "data" / "internal" / "h3_exports"
OUTPUT_DIR = ROOT / "data" / "internal" / "h3_normalized"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def normalize_header(value: str) -> str:
    return re.sub(r"[\s:：()（）/\\_\-]+", "", clean_text(value)).lower()


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
                return list(csv.DictReader(f, dialect=dialect))
        except (UnicodeDecodeError, csv.Error):
            continue
    raise RuntimeError(f"Cannot read CSV: {path}")


def find_header_row(rows: list[tuple[Any, ...]], aliases: set[str]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:30]):
        values = [normalize_header(clean_text(cell)) for cell in row]
        score = sum(1 for value in values if value in aliases)
        non_empty = sum(1 for value in values if value)
        if score > best_score or (score == best_score and non_empty > 2):
            best_score = score
            best_index = index
    return best_index


def read_xlsx_rows(path: Path, all_aliases: set[str]) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                continue
            header_index = find_header_row(values, all_aliases)
            headers = [clean_text(cell) for cell in values[header_index]]
            if not any(headers):
                continue
            for row in values[header_index + 1 :]:
                record = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers)) if headers[i]}
                if any(clean_text(value) for value in record.values()):
                    record["_source_sheet"] = sheet.title
                    records.append(record)
    finally:
        workbook.close()
    return records


def load_rows(path: Path, all_aliases: set[str]) -> list[dict[str, Any]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path, all_aliases)
    if path.suffix.lower() == ".csv":
        return read_csv_rows(path)
    return []


def alias_lookup(headers: list[str]) -> dict[str, str]:
    return {normalize_header(header): header for header in headers}


def score_table(path: Path, headers: list[str], table: dict[str, Any]) -> int:
    normalized_headers = set(normalize_header(header) for header in headers)
    score = 0
    for keyword in table.get("filename_keywords", []):
        if keyword.lower() in path.name.lower():
            score += 8
    for required in table.get("required_any", []):
        if any(normalize_header(required) in header for header in normalized_headers):
            score += 10
    for aliases in table.get("fields", {}).values():
        if any(normalize_header(alias) in normalized_headers for alias in aliases):
            score += 2
    return score


def choose_table(path: Path, rows: list[dict[str, Any]], tables: list[dict[str, Any]]) -> tuple[str, int]:
    headers = list(rows[0].keys()) if rows else []
    scored = [(table["id"], score_table(path, headers, table)) for table in tables]
    scored.sort(key=lambda item: item[1], reverse=True)
    if not scored or scored[0][1] < 8:
        return "", 0
    return scored[0]


def normalize_rows(path: Path, rows: list[dict[str, Any]], table: dict[str, Any]) -> list[dict[str, Any]]:
    if not rows:
        return []
    headers = list(rows[0].keys())
    lookup = alias_lookup(headers)
    normalized: list[dict[str, Any]] = []
    imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in rows:
        output = {
            "source_file": str(path),
            "source_sheet": clean_text(row.get("_source_sheet", "")),
            "imported_at": imported_at,
        }
        for canonical, aliases in table.get("fields", {}).items():
            source_header = ""
            for alias in aliases:
                key = normalize_header(alias)
                if key in lookup:
                    source_header = lookup[key]
                    break
            output[canonical] = clean_text(row.get(source_header, "")) if source_header else ""
        if any(output.get(field) for field in table.get("fields", {}).keys()):
            normalized.append(output)
    return normalized


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_status(status: dict[str, Any], table_rows: dict[str, list[dict[str, Any]]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "import_status.json").write_text(
        json.dumps(status, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    lines = [
        "# 美途能进销存系统导入状态",
        "",
        f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 导出目录：{EXPORT_DIR}",
        f"- 扫描文件：{status['scanned_files']}",
        f"- 已识别文件：{status['matched_files']}",
        f"- 未识别文件：{status['unmatched_files']}",
        "",
        "## 表数据行数",
        "",
        "| 标准表 | 行数 |",
        "| --- | ---: |",
    ]
    for table_id, rows in sorted(table_rows.items()):
        lines.append(f"| {table_id} | {len(rows)} |")
    if status["unmatched"]:
        lines.extend(["", "## 未识别文件", ""])
        lines.extend(f"- {item['file']}：{item['reason']}" for item in status["unmatched"])
    (OUTPUT_DIR / "import_status.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    tables = config.get("tables", [])
    all_aliases = {
        normalize_header(alias)
        for table in tables
        for aliases in table.get("fields", {}).values()
        for alias in aliases
    }
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    files = sorted(
        path
        for path in EXPORT_DIR.rglob("*")
        if path.is_file() and path.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
    )
    table_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    matched: list[dict[str, Any]] = []
    unmatched: list[dict[str, str]] = []

    for path in files:
        try:
            rows = load_rows(path, all_aliases)
            table_id, score = choose_table(path, rows, tables)
            if not table_id:
                unmatched.append({"file": str(path), "reason": "无法匹配到进销存标准表"})
                continue
            table = next(table for table in tables if table["id"] == table_id)
            normalized = normalize_rows(path, rows, table)
            table_rows[table_id].extend(normalized)
            matched.append({"file": str(path), "table": table_id, "score": score, "rows": len(normalized)})
        except Exception as exc:  # noqa: BLE001
            unmatched.append({"file": str(path), "reason": str(exc)})

    for table_id, rows in table_rows.items():
        write_csv(OUTPUT_DIR / f"{table_id}.csv", rows)

    status = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "primary_system": config.get("primary_system"),
        "scanned_files": len(files),
        "matched_files": len(matched),
        "unmatched_files": len(unmatched),
        "matched": matched,
        "unmatched": unmatched,
    }
    write_status(status, table_rows)
    print(f"Scanned files: {len(files)}")
    print(f"Matched files: {len(matched)}")
    print(f"Output: {OUTPUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
