from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DISCOVERY = ROOT / "data" / "internal" / "discovery"


DEFAULT_CATEGORIES = ["inventory", "bom", "purchase", "supplier", "sales", "finance"]
WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Profile high-value internal workbook candidates.")
    parser.add_argument(
        "--candidate-csv",
        default=str(DEFAULT_DISCOVERY / "candidate_files.csv"),
        help="Candidate file CSV from discover_internal_sources.py.",
    )
    parser.add_argument(
        "--out-dir",
        default=str(DEFAULT_DISCOVERY),
        help="Output directory.",
    )
    parser.add_argument(
        "--per-category",
        type=int,
        default=12,
        help="Number of top workbooks to profile per category.",
    )
    parser.add_argument(
        "--categories",
        nargs="*",
        default=DEFAULT_CATEGORIES,
        help="Categories to profile.",
    )
    return parser.parse_args()


def load_candidates(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def choose_workbooks(
    rows: list[dict[str, str]], categories: list[str], per_category: int
) -> list[dict[str, str]]:
    selected: dict[str, dict[str, str]] = {}
    by_category: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        ext = row.get("extension", "").lower()
        if ext not in WORKBOOK_EXTENSIONS:
            continue
        path = row.get("full_path", "")
        if "~$" in Path(path).name:
            continue
        for category in row.get("categories", "").split("|"):
            if category in categories:
                by_category[category].append(row)
    for category in categories:
        category_rows = sorted(
            by_category.get(category, []),
            key=lambda row: (int(row.get("score") or 0), row.get("modified_at", "")),
            reverse=True,
        )
        for row in category_rows[:per_category]:
            selected[row["full_path"]] = row
    return list(selected.values())


def compact_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\n", " ")
    if len(text) > 80:
        return text[:77] + "..."
    return text


def profile_xlsx(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets[:12]:
        preview_rows: list[list[str]] = []
        best_row_index = 0
        best_row_values: list[str] = []
        best_non_empty = 0
        for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            values = [compact_value(value) for value in row[:30]]
            non_empty = sum(1 for value in values if value)
            if row_index <= 8:
                preview_rows.append(values)
            if non_empty > best_non_empty:
                best_non_empty = non_empty
                best_row_index = row_index
                best_row_values = values
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "header_candidate_row": best_row_index,
                "header_candidate": best_row_values,
                "preview_rows": preview_rows,
            }
        )
    wb.close()
    return {"status": "ok", "sheets": sheets}


def profile_workbooks(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    profiles: list[dict[str, Any]] = []
    for row in rows:
        path = Path(row["full_path"])
        profile: dict[str, Any] = {
            "path": str(path),
            "name": path.name,
            "modified_at": row.get("modified_at", ""),
            "categories": row.get("categories", ""),
            "score": int(row.get("score") or 0),
            "extension": path.suffix.lower(),
        }
        try:
            if path.suffix.lower() in WORKBOOK_EXTENSIONS:
                profile.update(profile_xlsx(path))
            else:
                profile.update({"status": "skipped", "error": "unsupported extension"})
        except Exception as error:  # noqa: BLE001 - profile must continue after bad files.
            profile.update({"status": "error", "error": str(error)})
        profiles.append(profile)
    return profiles


def write_outputs(out_dir: Path, profiles: list[dict[str, Any]]) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "workbook_profiles.json").write_text(
        json.dumps(
            {
                "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "workbook_count": len(profiles),
                "profiles": profiles,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    lines = [
        "# 内部工作簿结构画像",
        "",
        f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 工作簿数量：{len(profiles)}",
        "",
    ]
    for profile in profiles:
        lines.append(f"## {profile['name']}")
        lines.append("")
        lines.append(f"- 路径：{profile['path']}")
        lines.append(f"- 分类：{profile['categories']}")
        lines.append(f"- 状态：{profile['status']}")
        if profile.get("error"):
            lines.append(f"- 错误：{profile['error']}")
        for sheet in profile.get("sheets", [])[:8]:
            headers = " | ".join(value for value in sheet.get("header_candidate", []) if value)
            if len(headers) > 260:
                headers = headers[:257] + "..."
            lines.append(
                f"- 表：{sheet['name']}，规模：{sheet['max_row']}行 x {sheet['max_column']}列，"
                f"表头候选第{sheet['header_candidate_row']}行：{headers}"
            )
        lines.append("")
    (out_dir / "workbook_profiles.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    args = parse_args()
    candidates = load_candidates(Path(args.candidate_csv))
    workbooks = choose_workbooks(candidates, args.categories, args.per_category)
    profiles = profile_workbooks(workbooks)
    write_outputs(Path(args.out_dir), profiles)
    print(f"Profiled workbooks: {len(profiles)}")
    print(f"Output: {args.out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
