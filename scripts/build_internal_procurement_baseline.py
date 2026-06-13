from __future__ import annotations

import csv
import json
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
DISCOVERY = ROOT / "data" / "internal" / "discovery"
OUTPUT = ROOT / "data" / "internal" / "procurement"
LATEST_JSON = ROOT / "data" / "latest.json"


REQUIRED_INVENTORY_HEADERS = ["物料名称", "当前库存数量", "日均出库数量", "预计可销售天数"]


HEADER_ALIASES = {
    "material_type": ["物料类型", "类型"],
    "material_name": ["物料名称", "名称", "产品名称"],
    "supplier": ["供应商"],
    "current_stock_qty": ["当前库存数量", "库存数量"],
    "inbound_qty": ["采购未到货数量", "未到货数量", "在途数量"],
    "near_arrival_qty": ["年前能到货数量", "3月能到货数量", "预计到货数量"],
    "buildable_qty": ["合计可生产数量", "可生产数量"],
    "finished_stock_qty": ["成品库存数量"],
    "total_available_qty": ["可生产+成品总量", "可生产数量+成品数量", "可生产+成品数量"],
    "outbound_30d_qty": ["近30天出库数量", "近30天出库数"],
    "daily_outbound_qty": ["日均出库数量"],
    "stock_days": ["预计可销售天数"],
    "warning_level": ["库存预警级别"],
    "suggested_order_qty": ["预采购下单数量", "预采购下单数"],
    "unit_price": ["单价"],
    "note": ["备注"],
}


COMMODITY_RULES = [
    ("battery", ["电芯", "电池", "磷酸铁锂", "电池包"], ["lithium_carbonate", "lfp_cathode_proxy"]),
    ("pcba", ["PCB", "PCBA", "线路板", "电路板", "主板", "灯板", "继电器板"], ["copper_foil_proxy", "fiberglass_cloth_proxy", "epoxy_resin", "tin"]),
    ("copper", ["铜", "线材", "导线", "夹子", "启动夹", "鳄鱼夹"], ["copper"]),
    ("tin", ["锡", "焊"], ["tin", "solder_tin_proxy"]),
    ("aluminum", ["铝", "散热"], ["aluminum"]),
    ("plastic", ["外壳", "塑胶", "塑料", "ABS", "PC", "PP"], ["abs", "pc", "pp"]),
    ("silicone", ["硅胶", "硅胶线", "硅"], ["organic_silicon_dmc", "industrial_silicon"]),
    ("pump", ["气泵", "充气", "电机", "马达"], ["copper", "aluminum", "steel_hc"]),
    ("hose", ["气管", "软管", "PVC"], ["pvc", "natural_rubber"]),
    ("paper", ["纸箱", "彩盒", "包装", "说明书", "纸"], ["corrugated_paper", "paper_pulp", "waste_paper"]),
    ("steel", ["钢", "弹簧", "螺丝", "五金"], ["steel_hc", "iron_ore"]),
]


@dataclass
class SourceWorkbook:
    path: Path
    sheet_name: str
    header_row: int


def load_market_risks() -> dict[str, dict[str, Any]]:
    data = json.loads(LATEST_JSON.read_text(encoding="utf-8"))
    return {item["material_id"]: item for item in data.get("latest", [])}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value):
            return None
        return float(value)
    text = str(value).replace(",", "").replace("%", "").strip()
    if text in {"", "-", "—", "None", "nan"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group())


def load_profiles() -> list[dict[str, Any]]:
    path = DISCOVERY / "workbook_profiles.json"
    data = json.loads(path.read_text(encoding="utf-8"))
    return data.get("profiles", [])


def sheet_date_key(name: str, index: int) -> tuple[int, int]:
    match = re.search(r"(\d{2})[-_.](\d{1,2})[-_.](\d{1,2})", name)
    if not match:
        return (0, index)
    year, month, day = (int(part) for part in match.groups())
    return (2000 + year) * 10000 + month * 100 + day, index


def find_inventory_source() -> SourceWorkbook:
    profiles = load_profiles()
    candidates: list[tuple[tuple[int, int], dict[str, Any], dict[str, Any]]] = []
    for profile in profiles:
        if profile.get("status") != "ok":
            continue
        if "inventory" not in profile.get("categories", ""):
            continue
        for index, sheet in enumerate(profile.get("sheets", [])):
            headers = " ".join(sheet.get("header_candidate") or [])
            if all(header in headers for header in REQUIRED_INVENTORY_HEADERS):
                candidates.append((sheet_date_key(sheet.get("name", ""), index), profile, sheet))
    if not candidates:
        raise RuntimeError("No inventory workbook with required headers found.")
    candidates.sort(key=lambda item: (item[0], item[1].get("modified_at", "")), reverse=True)
    _, profile, sheet = candidates[0]
    return SourceWorkbook(
        path=Path(profile["path"]),
        sheet_name=sheet["name"],
        header_row=int(sheet["header_candidate_row"]),
    )


def map_headers(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for normalized, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            for index, header in enumerate(headers):
                if alias == header or alias in header:
                    mapping[normalized] = index
                    break
            if normalized in mapping:
                break
    return mapping


def get_cell(row: tuple[Any, ...], mapping: dict[str, int], key: str) -> Any:
    index = mapping.get(key)
    if index is None or index >= len(row):
        return None
    return row[index]


def map_commodity(material_name: str, material_type: str, risks: dict[str, dict[str, Any]]) -> dict[str, Any]:
    haystack = f"{material_type} {material_name}"
    matched_ids: list[str] = []
    matched_group = "unmapped"
    for group, keywords, commodity_ids in COMMODITY_RULES:
        if any(keyword.lower() in haystack.lower() for keyword in keywords):
            matched_group = group
            matched_ids = [commodity_id for commodity_id in commodity_ids if commodity_id in risks]
            break
    if not matched_ids:
        return {
            "commodity_group": matched_group,
            "commodity_ids": "",
            "commodity_names": "",
            "market_up_probability": 0,
            "market_risk_level": "未映射",
            "market_trend": "",
        }
    matched = [risks[item_id] for item_id in matched_ids]
    top = max(matched, key=lambda item: item.get("up_probability") or 0)
    return {
        "commodity_group": matched_group,
        "commodity_ids": "|".join(matched_ids),
        "commodity_names": "|".join(item.get("material_name", "") for item in matched),
        "market_up_probability": top.get("up_probability") or 0,
        "market_risk_level": top.get("risk_level") or "",
        "market_trend": top.get("trend") or "",
    }


def inventory_risk_score(stock_days: float | None) -> int:
    if stock_days is None:
        return 55
    if stock_days <= 15:
        return 100
    if stock_days <= 30:
        return 85
    if stock_days <= 60:
        return 65
    if stock_days <= 90:
        return 45
    return 20


def warning_score(warning: str) -> int:
    if any(word in warning for word in ["红", "高", "紧急", "严重"]):
        return 90
    if any(word in warning for word in ["橙", "中", "预警"]):
        return 65
    if any(word in warning for word in ["黄", "低"]):
        return 45
    return 35 if warning else 30


def action_for(priority: float, stock_days: float | None, market_probability: float, suggested_qty: float | None) -> str:
    qty_positive = suggested_qty is not None and suggested_qty > 0
    if stock_days is not None and stock_days <= 30 and market_probability >= 55:
        return "立即补货并同步询价"
    if stock_days is not None and stock_days <= 30:
        return "立即复核库存并下单"
    if qty_positive and priority >= 65:
        return "本周分批补货"
    if market_probability >= 65:
        return "锁价谈判/供应商报价复核"
    if stock_days is not None and stock_days >= 90 and market_probability < 50:
        return "暂缓采购，压价跟踪"
    if priority >= 50:
        return "保持关注，滚动复核"
    return "暂不动作"


def read_inventory_rows(source: SourceWorkbook, risks: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(source.path, read_only=True, data_only=True)
    worksheet = workbook[source.sheet_name]
    header_values = [
        clean_text(cell)
        for cell in next(
            worksheet.iter_rows(min_row=source.header_row, max_row=source.header_row, values_only=True)
        )
    ]
    mapping = map_headers(header_values)
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=source.header_row + 1, values_only=True):
        material_name = clean_text(get_cell(row, mapping, "material_name"))
        if not material_name:
            continue
        if material_name in {"合计", "总计"}:
            continue
        material_type = clean_text(get_cell(row, mapping, "material_type"))
        supplier = clean_text(get_cell(row, mapping, "supplier"))
        stock_days = clean_number(get_cell(row, mapping, "stock_days"))
        daily_outbound = clean_number(get_cell(row, mapping, "daily_outbound_qty"))
        current_stock = clean_number(get_cell(row, mapping, "current_stock_qty"))
        inbound_qty = clean_number(get_cell(row, mapping, "inbound_qty"))
        suggested_qty = clean_number(get_cell(row, mapping, "suggested_order_qty"))
        unit_price = clean_number(get_cell(row, mapping, "unit_price"))
        warning = clean_text(get_cell(row, mapping, "warning_level"))
        commodity = map_commodity(material_name, material_type, risks)

        inv_score = inventory_risk_score(stock_days)
        market_score = float(commodity["market_up_probability"] or 0)
        warn_score = warning_score(warning)
        budget = (suggested_qty or 0) * (unit_price or 0)
        priority = round(inv_score * 0.42 + market_score * 0.28 + warn_score * 0.2 + (10 if budget > 0 else 0), 1)

        rows.append(
            {
                "snapshot_sheet": source.sheet_name,
                "material_type": material_type,
                "material_name": material_name,
                "supplier": supplier,
                "current_stock_qty": current_stock,
                "inbound_qty": inbound_qty,
                "buildable_qty": clean_number(get_cell(row, mapping, "buildable_qty")),
                "finished_stock_qty": clean_number(get_cell(row, mapping, "finished_stock_qty")),
                "total_available_qty": clean_number(get_cell(row, mapping, "total_available_qty")),
                "outbound_30d_qty": clean_number(get_cell(row, mapping, "outbound_30d_qty")),
                "daily_outbound_qty": daily_outbound,
                "stock_days": stock_days,
                "warning_level": warning,
                "suggested_order_qty": suggested_qty,
                "unit_price": unit_price,
                "suggested_budget": round(budget, 2),
                "commodity_group": commodity["commodity_group"],
                "commodity_ids": commodity["commodity_ids"],
                "commodity_names": commodity["commodity_names"],
                "market_up_probability": commodity["market_up_probability"],
                "market_risk_level": commodity["market_risk_level"],
                "market_trend": commodity["market_trend"],
                "purchase_priority": priority,
                "recommended_action": action_for(priority, stock_days, market_score, suggested_qty),
                "note": clean_text(get_cell(row, mapping, "note")),
            }
        )
    workbook.close()
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "snapshot_sheet",
        "material_type",
        "material_name",
        "supplier",
        "current_stock_qty",
        "inbound_qty",
        "buildable_qty",
        "finished_stock_qty",
        "total_available_qty",
        "outbound_30d_qty",
        "daily_outbound_qty",
        "stock_days",
        "warning_level",
        "suggested_order_qty",
        "unit_price",
        "suggested_budget",
        "commodity_group",
        "commodity_ids",
        "commodity_names",
        "market_up_probability",
        "market_risk_level",
        "market_trend",
        "purchase_priority",
        "recommended_action",
        "note",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_report(path: Path, source: SourceWorkbook, rows: list[dict[str, Any]]) -> None:
    sorted_rows = sorted(rows, key=lambda item: item["purchase_priority"], reverse=True)
    immediate = [row for row in sorted_rows if row["recommended_action"].startswith("立即")]
    split_buy = [row for row in sorted_rows if "分批" in row["recommended_action"]]
    lock_price = [row for row in sorted_rows if "锁价" in row["recommended_action"]]
    total_budget = sum(row.get("suggested_budget") or 0 for row in rows)
    mapped_count = sum(1 for row in rows if row.get("commodity_ids"))

    lines = [
        "# 内部采购建议基线",
        "",
        f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 来源工作簿：{source.path}",
        f"- 来源工作表：{source.sheet_name}",
        f"- 物料行数：{len(rows)}",
        f"- 已映射行情风险：{mapped_count}",
        f"- 预采购预算合计：{total_budget:,.2f}",
        f"- 立即处理：{len(immediate)}",
        f"- 本周分批补货：{len(split_buy)}",
        f"- 锁价/报价复核：{len(lock_price)}",
        "",
        "## 优先处理清单",
        "",
        "| 优先级 | 动作 | 物料 | 供应商 | 库存天数 | 建议采购 | 单价 | 预算 | 行情风险 | 代理指标 |",
        "| --- | --- | --- | --- | ---: | ---: | ---: | ---: | --- | --- |",
    ]
    for row in sorted_rows[:20]:
        lines.append(
            "| {priority} | {action} | {material} | {supplier} | {days} | {qty} | {price} | {budget} | {risk} {prob}% | {commodity} |".format(
                priority=row["purchase_priority"],
                action=row["recommended_action"],
                material=row["material_name"],
                supplier=row["supplier"],
                days="" if row["stock_days"] is None else f"{row['stock_days']:.1f}",
                qty="" if row["suggested_order_qty"] is None else f"{row['suggested_order_qty']:,.0f}",
                price="" if row["unit_price"] is None else f"{row['unit_price']:,.2f}",
                budget=f"{row['suggested_budget']:,.2f}",
                risk=row["market_risk_level"],
                prob=row["market_up_probability"],
                commodity=row["commodity_names"] or "未映射",
            )
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    risks = load_market_risks()
    source = find_inventory_source()
    rows = read_inventory_rows(source, risks)
    rows = sorted(rows, key=lambda item: item["purchase_priority"], reverse=True)
    write_csv(OUTPUT / "procurement_baseline.csv", rows)
    write_report(OUTPUT / "procurement_baseline.md", source, rows)
    print(f"Source workbook: {source.path}")
    print(f"Source sheet: {source.sheet_name}")
    print(f"Rows: {len(rows)}")
    print(f"Output: {OUTPUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
