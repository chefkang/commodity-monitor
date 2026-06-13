from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
DISCOVERY = ROOT / "data" / "internal" / "discovery"
OUTPUT = ROOT / "data" / "internal" / "procurement"


REQUIRED_HEADERS = ["SKU名称", "物料名称", "标准计价", "用量"]
OPTIONAL_HEADERS = ["BOM价格", "是/否有子物料"]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_bom_workbook() -> Path:
    candidates_path = DISCOVERY / "candidate_files.csv"
    with candidates_path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    preferred = [
        row
        for row in rows
        if row.get("extension") in {".xlsx", ".xlsm"}
        and "SKU组装单2" in row.get("name", "")
    ]
    if not preferred:
        preferred = [
            row
            for row in rows
            if row.get("extension") in {".xlsx", ".xlsm"}
            and "SKU组装单" in row.get("name", "")
        ]
    if not preferred:
        raise RuntimeError("No SKU BOM workbook found.")
    preferred.sort(key=lambda row: (row.get("modified_at", ""), int(row.get("score") or 0)), reverse=True)
    return Path(preferred[0]["full_path"])


def find_header(values: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for expected in REQUIRED_HEADERS + OPTIONAL_HEADERS:
        for index, value in enumerate(values):
            if expected == value or expected in value:
                mapping[expected] = index
                break
    return mapping


def extract_bom(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "总表" if "总表" in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]

    header_row = None
    header_map: dict[str, int] = {}
    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        values = [clean_text(value) for value in row]
        mapping = find_header(values)
        if all(header in mapping for header in REQUIRED_HEADERS):
            header_row = row_index
            header_map = mapping
            break
    if header_row is None:
        raise RuntimeError("Could not find BOM header row.")

    rows: list[dict[str, Any]] = []
    current_sku = ""
    current_bom_price: float | None = None
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        raw_sku = clean_text(row[header_map["SKU名称"]])
        if raw_sku:
            current_sku = raw_sku
        sku = current_sku
        if "BOM价格" in header_map:
            raw_bom_price = clean_number(row[header_map["BOM价格"]])
            if raw_bom_price is not None:
                current_bom_price = raw_bom_price
        material_name = clean_text(row[header_map["物料名称"]])
        if not sku and not material_name:
            continue
        if not material_name or material_name in {"合计", "总计"}:
            continue
        unit_price = clean_number(row[header_map["标准计价"]])
        qty = clean_number(row[header_map["用量"]])
        line_cost = (unit_price or 0) * (qty or 0)
        rows.append(
            {
                "source_workbook": str(path),
                "source_sheet": sheet_name,
                "sku": sku,
                "bom_price": current_bom_price,
                "material_name": material_name,
                "has_sub_material": clean_text(row[header_map.get("是/否有子物料", -1)])
                if "是/否有子物料" in header_map
                else "",
                "unit_price": unit_price,
                "qty_per_unit": qty,
                "line_cost": round(line_cost, 4),
            }
        )
    workbook.close()
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "source_workbook",
                "source_sheet",
                "sku",
                "bom_price",
                "material_name",
                "has_sub_material",
                "unit_price",
                "qty_per_unit",
                "line_cost",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)


def write_report(path: Path, rows: list[dict[str, Any]]) -> None:
    sku_counts = Counter(row["sku"] for row in rows if row["sku"])
    material_counts = Counter(row["material_name"] for row in rows if row["material_name"])
    sku_costs: dict[str, float] = {}
    for row in rows:
        sku = row["sku"]
        if not sku:
            continue
        sku_costs[sku] = sku_costs.get(sku, 0) + float(row.get("line_cost") or 0)

    lines = [
        "# 内部SKU BOM基线",
        "",
        f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- BOM行数：{len(rows)}",
        f"- SKU数量：{len(sku_counts)}",
        f"- 物料名称数量：{len(material_counts)}",
        "",
        "## SKU成本初算排行",
        "",
        "| SKU | BOM行数 | 初算BOM成本 |",
        "| --- | ---: | ---: |",
    ]
    for sku, cost in sorted(sku_costs.items(), key=lambda item: item[1], reverse=True)[:30]:
        lines.append(f"| {sku} | {sku_counts[sku]} | {cost:,.2f} |")
    lines.extend(["", "## 高频物料", "", "| 物料 | 出现次数 |", "| --- | ---: |"])
    for material, count in material_counts.most_common(30):
        lines.append(f"| {material} | {count} |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    path = find_bom_workbook()
    rows = extract_bom(path)
    write_csv(OUTPUT / "sku_bom_baseline.csv", rows)
    write_report(OUTPUT / "sku_bom_baseline.md", rows)
    print(f"Source workbook: {path}")
    print(f"Rows: {len(rows)}")
    print(f"Output: {OUTPUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
